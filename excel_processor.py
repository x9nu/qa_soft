import openpyxl

def load_questions_from_excel(file_path):
    # 容器准备存储所有问题
    questions = {
        '单选': [],
        '判断': []
    }

    try:
        # 加载Excel工作簿
        workbook = openpyxl.load_workbook(file_path)
        # ... (您原有的Excel文件处理逻辑)
        # 处理单选题
        if '单选' in workbook.sheetnames:
            single_choice_sheet = workbook['单选']
            for row in single_choice_sheet.iter_rows(min_row=3, values_only=True):
                question = {
                    '序号': row[0],
                    '题目': row[1],
                    'A': row[2],
                    'B': row[3],
                    'C': row[4],
                    'D': row[5],
                    '答案': row[6]
                }
                questions['单选'].append(question)

        # 处理判断题
        if '判断' in workbook.sheetnames:
            true_false_sheet = workbook['判断']
            for row in true_false_sheet.iter_rows(min_row=3, values_only=True):  # 从第三行开始读取数据
                question = {
                    '序号': row[0],
                    '题目': row[1],
                    '答案': bool(row[2])  # 假定1为True，0为False
                }
                questions['判断'].append(question)

    except Exception as e:
        print(f"An error occurred while loading the Excel file: {e}")

    return questions

# 如果需要测试此函数，可以取消注释以下代码
# if __name__ == "__main__":
#     path = "电竞1+X-中级-题库.xlsx"
#     all_questions = load_questions_from_excel(path)
#     print(all_questions)